import csv
from abc import abstractmethod, ABCMeta
import config as cfg
from openpyxl import load_workbook
import os
import hashlib
import base64


class ParserException(Exception):
    pass


class ParserFileNameException(Exception):
    pass


class ParserTypeNotFoundException(Exception):
    pass


class ParserHeaderNotValidException(Exception):
    pass


class Parser(object):

    def __init__(self, filename, from_path=cfg.FROM_PATH):
        self.tariff_file = '%s/%s' % (from_path, filename)

    @staticmethod
    def is_drug_tariff_part(tariff_period_category_row):
        return 'Drug Tariff Part' in tariff_period_category_row[0]

    @staticmethod
    def is_category_m_prices(tariff_period_category_row):
        return 'Category M Prices' in tariff_period_category_row[0]

    def get_parser(self):
        filename, file_extension = os.path.splitext(self.tariff_file)
        if file_extension == '.xlsx':
            tariff_period_category_row = self.get_first_line_from_xlsx()
        elif file_extension == '.csv':
            tariff_period_category_row = self.get_first_line_from_csv()
        else:
            raise ParserFileNameException('Unrecognized extension: %s' % self.tariff_file)

        if self.is_drug_tariff_part(tariff_period_category_row):
            return DrugPartMParser(self.tariff_file, file_extension)
        elif self.is_category_m_prices(tariff_period_category_row):
            return CategoryMParser(self.tariff_file, file_extension)
        else:
            raise ParserTypeNotFoundException('File %s row: %s' % (self.tariff_file, tariff_period_category_row))

    def get_first_line_from_csv(self):
        with open(self.tariff_file, 'r') as csvfile:
            csvreader = csv.reader(csvfile)
            tariff_period_category_row = next(csvreader)
            return tariff_period_category_row

    def get_first_line_from_xlsx(self):
        wb = load_workbook(filename=self.tariff_file, read_only=True)
        try:
            active_sheet = wb.active
            rows = ([str(cell.value) if cell.value else '' for cell in row] for row in active_sheet.rows)
            tariff_period_category_row = next(rows)
            return tariff_period_category_row
        finally:
            wb._archive.close()


class BasicParser(object):
    __metaclass__ = ABCMeta

    def __init__(self, tariff_file, file_extension='.csv'):
        self.tariff_file = tariff_file
        self.file_extension = file_extension

    def parse(self, file):
        if self.file_extension == '.csv':
            return self.parsecsv(file)
        elif self.file_extension == '.xlsx':
            return self.parsexls(file)
        else:
            raise ParserFileNameException('Unrecognized extension: %s' % self.tariff_file)

    def parsecsv(self, file):
        with open(self.tariff_file, 'r') as csvfile:
            csvreader = csv.reader(csvfile)
            tariff_period_category_row = next(csvreader)
            return self.tariff_to_json(tariff_period_category_row, self.get_category_period_extractor, csvreader, file)

    def parsexls(self, file):
        wb = load_workbook(filename=self.tariff_file, read_only=True)
        active_sheet = wb.active
        rows = ([str(cell.value) if cell.value else '' for cell in row] for row in active_sheet.rows)
        tariff_period_category_row = next(rows)
        return self.tariff_to_json(tariff_period_category_row, self.get_category_period_extractor, rows, file)

    def tariff_to_json(self, tariff_period_category_row, category_period_extractor, csvreader, file):
        blank_line = next(csvreader)
        empty_line = [not col for col in blank_line]
        if all(empty_line):
            raw_header = self.normalize_row(next(csvreader))
            header = self.normalize_header(raw_header)
            if self.is_header_valid(header):
                category, period = category_period_extractor(tariff_period_category_row)
                medicines = []
                for line in csvreader:
                    medicine = {
#                        'id': uuid.uuid4().hex,
                        'category': category,
                        'period': period,
                        'url': file['url'],
                        'filename': file['path']
                    }
                    line = self.normalize_row(line)
                    medicine.update({head if head else 'unit': val for head, val in zip(header, line)})
                    digest = self.get_digest(medicine)
                    medicine['digest'] = digest.decode('utf-8')
                    medicines.append(medicine)
                return medicines
            else:
                raise ParserHeaderNotValidException('Bad header file %s header: %s' % (self.tariff_file, header))
        else:
            raise ParserException("Drug tariff %s failed, should be a blank row: %s" % (self.tariff_file, blank_line))

    @abstractmethod
    def get_category_period_extractor(self, tariff_period_category):
        pass

    @abstractmethod
    def is_header_valid(self, header):
        pass

    @staticmethod
    def normalize_row(row):
        return [col.strip() for col in row]

    @staticmethod
    def normalize_header(raw_header):
        norm = {
            "Drug Name": "Medicine",
            "Pack size": "Pack Size",
            "Spec Cont Ind": "Special Container",
            "Special container": "Special Container",
            "Special container indicator": "Special Container",
            "Drug Tariff Category": "category"
        }
        normalized_header = [norm.get(h, h) for h in raw_header]
        return normalized_header

    @staticmethod
    def get_digest(doc):
        vals = [v for k, v in doc.items() if k not in ["_id", "digest", 'filename', 'url', 'dupes']]
        s_vals = "".join(vals)
        b_vals = bytearray(s_vals, "utf-8")
        algo = hashlib.sha3_256()
        algo.update(b_vals)
        digest = algo.digest()
        return base64.b64encode(digest)

    @staticmethod
    def digest_header(header):
        header.append('digest')
        return header

    @staticmethod
    def tolower_row(row):
        return [col.lower() for col in row]


class DrugPartMParser(BasicParser):

    def get_category_period_extractor(self, tariff_period_category):
        category_period = tariff_period_category[0].split(' ')
        category = ' '.join(category_period[1:])
        period = category_period[:1][0]
        return category, period

    def is_header_valid(self, header):
        lower_header = self.tolower_row(header)
        return 'Basic Price'.lower() in lower_header and 'Pack Size'.lower() in lower_header


class CategoryMParser(BasicParser):

    def get_category_period_extractor(self, tariff_period_category_row):
        category_period = tariff_period_category_row[0].split('-')
        category = category_period[:1][0]
        period = category_period[1:][0]
        return category, period

    def is_header_valid(self, header):
        lower_header = self.tolower_row(header)
        return 'Basic Price'.lower() in lower_header and 'Pack Size'.lower() in lower_header
